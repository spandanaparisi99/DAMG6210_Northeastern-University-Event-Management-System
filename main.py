import datetime
import pandas as pd
import tweepy
import pymysql
import config
import MySQLdb
import mysql.connector
pymysql.install_as_MySQLdb()
import openpyxl as xl
import xlwt
from datetime import date
import xlrd
import random

API_KEY = config.API_KEY
API_KEY_SECRET = config.API_KEY_SECRET
API_TOKEN = config.API_TOKEN
API_TOKEN_SECRET = config.API_TOKEN_SECRET
BEARER_TOKEN = config.BEARER_TOKEN

USER = config.USER
PASSWORD = config.PASSWORD
HOST = config.HOST
DATABASE = config.DATABASE
PORT = config.PORT

auth = tweepy.OAuthHandler(API_KEY, API_KEY_SECRET)
auth.set_access_token(API_TOKEN, API_TOKEN_SECRET)
api = tweepy.API(auth)


print("Enter the hashtag to search for")
word_input = input()

print("Enter the no. of tweets to be scrapped:")
tweet_input = input()
print("tweets fetching ...")
print("for user table")

limit = 10

tweets = tweepy.Cursor(api.search_tweets, q=word_input, count=tweet_input, tweet_mode='extended').items(limit)

columns = ['twitter_handle', 'user_name', 'follower_count', 'following_count', 'join_date']
data = []

for tweet in tweets:
    data.append([tweet.user.screen_name,tweet.user.screen_name, random.randint(0,1000), random.randint(0,1000), '2008-11-11'])

# Creating a dataframe
df = pd.DataFrame(data, columns=columns)

print(df)

# Writing Dataframe to excel
df.to_excel('user.xls', index=False, header=True)

# Open the workbook and define the worksheet
book = xlrd.open_workbook("user.xls")
sheet = book.sheet_by_name("Sheet1")


#create connection
con = mysql.connector.connect(
  host=HOST,
  user=USER,
  password=PASSWORD,
  database=DATABASE
)

#create cursor
cursor = con.cursor()

query = """INSERT INTO user (twitter_handle, user_name , follower_count, following_count, join_date) VALUES (%s, %s, %s, %s, %s)"""

# Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
for r in range(1, sheet.nrows):
    twitter_handle = sheet.cell(r,0).value
    user_name = sheet.cell(r,1).value
    follower_count = sheet.cell(r,2).value
    following_count	= sheet.cell(r,3).value
    join_date = sheet.cell(r,4).value

    # Assign values from each row
    values = (twitter_handle, user_name, follower_count, following_count, join_date)

    # Execute sql Query
    cursor.execute(query, values)

# Commit
con.commit()


print("Enter the hashtag to search for")
word_input2 = input()

print("Enter the no. of tweets to be scrapped:")
tweet_input2 = input()
print("tweets fetching ...")
print("for tweets table")

limit2 = 10

tweets2 = tweepy.Cursor(api.search_tweets, q=word_input2, count=tweet_input2, tweet_mode='extended').items(limit2)

columns2 = ['tweet_id', 'twitter_handle', 'tweet_text', 'Created at', 'hashtag']
data2 = []

for tweet in tweets2:
    data2.append([tweet.id,tweet.user.screen_name, tweet.full_text,'2008-11-11', word_input2])

# Creating a dataframe
df2 = pd.DataFrame(data2, columns=columns2)

print(df2)

# Writing Dataframe to excel
df2.to_excel('tweets.xls', index=False, header=True)

# Open the workbook and define the worksheet
book2 = xlrd.open_workbook("tweets.xls")
sheet2 = book2.sheet_by_name("Sheet1")


query2 = """INSERT INTO tweets (tweet_id, twitter_handle , tweet_text, Created_at, hashtags) VALUES (%s, %s, %s, %s, %s)"""

# Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
for r in range(1, sheet2.nrows):
    tweet_id = sheet2.cell(r,0).value
    twitter_hdle = sheet2.cell(r,1).value
    tweet_text = sheet2.cell(r,2).value
    created_at	= sheet2.cell(r,3).value
    hashtag = sheet2.cell(r,4).value

    # Assign values from each row
    values2 = (tweet_id, twitter_hdle, tweet_text, created_at, hashtag)

    # Execute sql Query
    cursor.execute(query2, values2)

# Commit
con.commit()
cursor.close()
con.close()


